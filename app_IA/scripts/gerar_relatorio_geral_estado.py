from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Dict, Iterable, List, Sequence


@dataclass
class Paths:
    data_root: Path
    new_date_root: Path
    analyse_root: Path
    app_root: Path
    referencia_root: Path
    prompts_root: Path
    coleta_root: Path
    pdf_root: Path
    resposta_root: Path
    texto_root: Path
    validacao_root: Path
    app_data_root: Path
    consolidado_csv: Path
    validacao_controle_csv: Path
    base_docs_csv: Path
    vinculos_csv: Path
    mapeamento_csv: Path


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def discover_paths(script_path: Path) -> Paths:
    app_root = script_path.resolve().parent.parent
    analyse_root = app_root.parent
    new_date_root = analyse_root.parent
    data_root = new_date_root.parent

    return Paths(
        data_root=data_root,
        new_date_root=new_date_root,
        analyse_root=analyse_root,
        app_root=app_root,
        referencia_root=new_date_root / "dados_iniciais_ref" / "data_2_estrutura",
        prompts_root=new_date_root / "prompts_individuais_7em7",
        coleta_root=new_date_root / "coleta_md_por_estado",
        pdf_root=new_date_root / "pdfs_baixados",
        resposta_root=analyse_root / "Resposta_csv",
        texto_root=analyse_root / "Texto_Qualitativo",
        validacao_root=analyse_root / "Validação_Manual",
        app_data_root=app_root / "data",
        consolidado_csv=analyse_root / "Resposta_csv" / "analise_notebooklm_consolidado.csv",
        validacao_controle_csv=analyse_root / "Validação_Manual" / "validacao_controle.csv",
        base_docs_csv=app_root / "data" / "base_documentos.csv",
        vinculos_csv=app_root / "data" / "vinculos_pdf_campus.csv",
        mapeamento_csv=new_date_root / "output_csv" / "mapeamento_ppc_polos.csv",
    )


def read_csv_dict(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_ref_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []

    rows: List[Dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not row or not any((cell or "").strip() for cell in row):
                continue

            items = list(row) + [""] * 4
            rows.append(
                {
                    "INSTITUICAO": items[0].strip(),
                    "MODALIDADE": items[1].strip(),
                    "NOME_POLO_OU_CAMPUS": items[2].strip(),
                    "MUNICIPIO_POLO": items[3].strip(),
                }
            )
    return rows


def count_files(path: Path, pattern: str, recursive: bool = True) -> int:
    if not path.exists():
        return 0
    if recursive:
        return sum(1 for _ in path.rglob(pattern))
    return sum(1 for _ in path.glob(pattern))


def normalize_doc_id(value: str) -> str:
    text = (value or "").strip().upper()
    if text.startswith("DOC"):
        digits = "".join(ch for ch in text if ch.isdigit())
        if digits:
            return f"DOC{int(digits):04d}"
    return text


def normalize_filename(value: str) -> str:
    return (value or "").strip().lower()


def build_analysis_token(doc_id: str, nome_arquivo_local: str) -> str:
    doc_norm = normalize_doc_id(doc_id)
    filename_norm = normalize_filename(nome_arquivo_local)

    if doc_norm and filename_norm:
        return f"{doc_norm}||{filename_norm}"
    return doc_norm or filename_norm


def analysis_token_from_row(row: Dict[str, str]) -> str:
    return build_analysis_token(row.get("DOC_ID", ""), row.get("NOME_ARQUIVO_LOCAL", ""))


def format_analysis_label(doc_id: str, nome_arquivo_local: str) -> str:
    doc_norm = normalize_doc_id(doc_id)
    filename = (nome_arquivo_local or "").strip()
    if doc_norm and filename:
        return f"{doc_norm} | {filename}"
    return doc_norm or filename or "SEM_IDENTIFICADOR"


def format_analysis_label_from_token(token: str) -> str:
    if "||" not in token:
        return token or "SEM_IDENTIFICADOR"
    doc_id, filename_norm = token.split("||", 1)
    if doc_id and filename_norm:
        return f"{doc_id} | {filename_norm}"
    return token


def parse_int(value: str, default: int = 0) -> int:
    text = (value or "").strip()
    if not text:
        return default
    if text.isdigit():
        return int(text)
    try:
        return int(float(text.replace(",", ".")))
    except Exception:
        return default


def unique_latest_rows(rows: Sequence[Dict[str, str]]) -> List[Dict[str, str]]:
    latest: Dict[str, Dict[str, str]] = {}

    for row in rows:
        token = analysis_token_from_row(row)
        if not token:
            continue

        prev = latest.get(token)
        if prev is None:
            latest[token] = dict(row)
            continue

        prev_ts = (prev.get("DATA_REGISTRO_UTC", "") or "").strip()
        cur_ts = (row.get("DATA_REGISTRO_UTC", "") or "").strip()
        if cur_ts >= prev_ts:
            latest[token] = dict(row)

    return sorted(
        latest.values(),
        key=lambda item: (
            normalize_doc_id(item.get("DOC_ID", "")),
            normalize_filename(item.get("NOME_ARQUIVO_LOCAL", "")),
        ),
    )


def distribution(rows: Sequence[Dict[str, str]], key: str) -> Dict[str, int]:
    counter = Counter()
    for row in rows:
        value = (row.get(key, "") or "").strip()
        counter[value if value else "VAZIO"] += 1
    return dict(counter)


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    if not headers:
        return ""

    out = []
    out.append("| " + " | ".join(headers) + " |")
    out.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        out.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(out)


def maybe_xlsx_stats(path: Path) -> Dict[str, str]:
    data: Dict[str, str] = {
        "path": str(path),
        "exists": "1" if path.exists() else "0",
        "workbook_rows": "",
        "workbook_cols": "",
        "sheet_name": "",
        "parser": "",
        "error": "",
    }

    if not path.exists():
        return data

    try:
        import openpyxl  # type: ignore
        from openpyxl.utils.cell import range_boundaries  # type: ignore

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row is None or max_col is None:
            try:
                dim = ws.calculate_dimension()
                min_col, min_row, max_col_dim, max_row_dim = range_boundaries(dim)
                max_row = max_row if max_row is not None else max_row_dim
                max_col = max_col if max_col is not None else max_col_dim
            except Exception:
                pass

        data["sheet_name"] = str(ws.title)
        data["workbook_rows"] = str(max_row if max_row is not None else "NA")
        data["workbook_cols"] = str(max_col if max_col is not None else "NA")
        data["parser"] = "openpyxl"
        wb.close()
    except Exception as exc:
        data["parser"] = "unavailable"
        data["error"] = str(exc)

    return data


def compute_state_overview(paths: Paths, uf: str) -> Dict[str, object]:
    uf = uf.upper()

    ref_path = paths.referencia_root / f"Pesquisa_{uf}.csv"
    ref_rows = read_ref_rows(ref_path)

    prompts_count = count_files(paths.prompts_root / uf, "*.md")
    coleta_count = count_files(paths.coleta_root / uf, "*.md")
    pdf_count = count_files(paths.pdf_root / uf, "*.pdf", recursive=False)

    texto_count = count_files(paths.texto_root / uf, "*.md")
    validacao_md_count = count_files(paths.validacao_root / uf, "*.md")

    consolidated_rows = read_csv_dict(paths.consolidado_csv)
    rows_uf = [row for row in consolidated_rows if (row.get("UF", "") or "").upper() == uf]
    rows_uf_latest = unique_latest_rows(rows_uf)

    base_rows = [row for row in read_csv_dict(paths.base_docs_csv) if (row.get("UF", "") or "").upper() == uf]
    vinculos_rows = [row for row in read_csv_dict(paths.vinculos_csv) if (row.get("UF", "") or "").upper() == uf]

    ref_instituicoes = distribution(ref_rows, "INSTITUICAO")
    ref_modalidades = distribution(ref_rows, "MODALIDADE")

    validacao_status = distribution(rows_uf_latest, "STATUS_VALIDACAO_HUMANA")
    comparacao_status = distribution(rows_uf_latest, "STATUS_COMPARACAO_BASE")

    tdic_present = sum(1 for row in rows_uf_latest if (row.get("PRESENCA_TDIC_0_1", "") or "").strip() == "1")
    ia_present = sum(1 for row in rows_uf_latest if (row.get("PRESENCA_IA_0_1", "") or "").strip() == "1")

    nivel_integracao_values = [
        parse_int(row.get("NIVEL_INTEGRACAO_0_1_2_3", ""), default=0) for row in rows_uf_latest
    ]
    nivel_integracao_media = round(mean(nivel_integracao_values), 2) if nivel_integracao_values else 0.0

    base_tokens: set[str] = set()
    analysed_tokens: set[str] = set()
    token_to_label: Dict[str, str] = {}

    for row in base_rows:
        token = analysis_token_from_row(row)
        if not token:
            continue
        base_tokens.add(token)
        token_to_label[token] = format_analysis_label(row.get("DOC_ID", ""), row.get("NOME_ARQUIVO_LOCAL", ""))

    for row in rows_uf_latest:
        token = analysis_token_from_row(row)
        if not token:
            continue
        analysed_tokens.add(token)
        token_to_label[token] = format_analysis_label(row.get("DOC_ID", ""), row.get("NOME_ARQUIVO_LOCAL", ""))

    docs_pending_tokens = sorted(base_tokens - analysed_tokens)
    docs_pending = [token_to_label.get(token, format_analysis_label_from_token(token)) for token in docs_pending_tokens]

    coverage_pct = round((len(analysed_tokens) / len(base_tokens)) * 100, 1) if base_tokens else 0.0

    incons_by_doc: Dict[str, Dict[str, object]] = {}
    grouped = defaultdict(list)
    for row in vinculos_rows:
        token = analysis_token_from_row(row)
        if token:
            grouped[token].append(row)

    total_vinculos = 0
    total_incons_ies = 0

    for token, rows in grouped.items():
        if not token:
            continue
        total = len(rows)
        incons = sum(1 for row in rows if (row.get("CONSISTENCIA_IES_0_1", "") or "").strip() == "0")
        total_vinculos += total
        total_incons_ies += incons
        incons_by_doc[token] = {
            "total_vinculos": total,
            "incons_ies": incons,
            "incons_ies_pct": round((incons / total) * 100, 1) if total else 0.0,
        }

    global_incons_pct = round((total_incons_ies / total_vinculos) * 100, 1) if total_vinculos else 0.0

    doc_rows = []
    for row in rows_uf_latest:
        doc_id = normalize_doc_id(row.get("DOC_ID", ""))
        token = analysis_token_from_row(row)
        doc_rows.append(
            {
                "DOC_KEY": token,
                "DOC_ID": doc_id,
                "NOME_ARQUIVO_LOCAL": row.get("NOME_ARQUIVO_LOCAL", ""),
                "STATUS_VALIDACAO_HUMANA": row.get("STATUS_VALIDACAO_HUMANA", ""),
                "STATUS_COMPARACAO_BASE": row.get("STATUS_COMPARACAO_BASE", ""),
                "PRESENCA_TDIC_0_1": row.get("PRESENCA_TDIC_0_1", ""),
                "PRESENCA_IA_0_1": row.get("PRESENCA_IA_0_1", ""),
                "NIVEL_INTEGRACAO_0_1_2_3": row.get("NIVEL_INTEGRACAO_0_1_2_3", ""),
                "NIVEL_CONFIANCA_CODIFICADOR_1_2_3": row.get("NIVEL_CONFIANCA_CODIFICADOR_1_2_3", ""),
                "QTD_VINCULOS_BASE": row.get("QTD_VINCULOS_BASE", ""),
            }
        )

    return {
        "uf": uf,
        "reference_csv_path": str(ref_path),
        "reference_rows": len(ref_rows),
        "reference_institutions": ref_instituicoes,
        "reference_modalities": ref_modalidades,
        "prompts_md_count": prompts_count,
        "coleta_md_count": coleta_count,
        "pdf_count": pdf_count,
        "texto_qualitativo_md_count": texto_count,
        "validacao_manual_md_count": validacao_md_count,
        "docs_analisados": len(rows_uf_latest),
        "docs_analisados_ids": sorted(
            [token_to_label.get(token, format_analysis_label_from_token(token)) for token in analysed_tokens]
        ),
        "docs_base": len(base_tokens),
        "docs_base_ids": sorted([token_to_label.get(token, format_analysis_label_from_token(token)) for token in base_tokens]),
        "docs_pending": docs_pending,
        "coverage_pct": coverage_pct,
        "validacao_status": validacao_status,
        "comparacao_status": comparacao_status,
        "tdic_present": tdic_present,
        "ia_present": ia_present,
        "nivel_integracao_media": nivel_integracao_media,
        "inconsistencia_vinculos_total": total_vinculos,
        "inconsistencia_vinculos_ies": total_incons_ies,
        "inconsistencia_vinculos_ies_pct": global_incons_pct,
        "inconsistencia_por_doc": incons_by_doc,
        "docs_detail": doc_rows,
    }


def list_state_counts(paths: Paths) -> List[Dict[str, object]]:
    states = []
    ref_files = sorted(paths.referencia_root.glob("Pesquisa_*.csv"))

    base_rows = read_csv_dict(paths.base_docs_csv)
    base_count_by_uf: Dict[str, int] = defaultdict(int)
    for row in base_rows:
        uf = (row.get("UF", "") or "").upper().strip()
        if uf:
            base_count_by_uf[uf] += 1

    consolidated_rows = read_csv_dict(paths.consolidado_csv)
    by_uf_latest: Dict[str, set[str]] = defaultdict(set)
    for row in unique_latest_rows(consolidated_rows):
        uf = (row.get("UF", "") or "").upper().strip()
        token = analysis_token_from_row(row)
        if uf and token:
            by_uf_latest[uf].add(token)

    for ref_file in ref_files:
        uf = ref_file.stem.replace("Pesquisa_", "").upper()
        ref_count = len(read_ref_rows(ref_file))
        pdf_count = count_files(paths.pdf_root / uf, "*.pdf", recursive=False)
        prompts_count = count_files(paths.prompts_root / uf, "*.md")
        coleta_count = count_files(paths.coleta_root / uf, "*.md")
        analysed_count = len(by_uf_latest.get(uf, set()))
        docs_base_app = int(base_count_by_uf.get(uf, 0))
        universe_for_pending = docs_base_app if docs_base_app > 0 else pdf_count
        pending_pdf = max(universe_for_pending - analysed_count, 0)

        states.append(
            {
                "UF": uf,
                "REF_ROWS": ref_count,
                "PROMPTS_MD": prompts_count,
                "COLETA_MD": coleta_count,
                "PDFS_BRUTOS": pdf_count,
                "DOCS_BASE_APP": docs_base_app,
                "ANALISADOS": analysed_count,
                "PENDENTES_ANALISE": pending_pdf,
                "COBERTURA_ANALISE_PDFS_PCT": round((analysed_count / pdf_count) * 100, 1) if pdf_count else 0.0,
                "COBERTURA_ANALISE_BASE_APP_PCT": (
                    round((analysed_count / docs_base_app) * 100, 1) if docs_base_app else None
                ),
            }
        )

    return sorted(states, key=lambda item: item["UF"])


def recommend_next_state(states: Sequence[Dict[str, object]], current_uf: str) -> Dict[str, object] | None:
    candidates = [
        item
        for item in states
        if item["UF"] != current_uf and int(item["PDFS_BRUTOS"]) > 0 and int(item["PENDENTES_ANALISE"]) > 0
    ]

    if not candidates:
        return None

    # Prefer state with more PDFs ready and more pending analysis.
    candidates.sort(
        key=lambda item: (
            1 if int(item.get("DOCS_BASE_APP", 0)) > 0 else 0,
            int(item["PENDENTES_ANALISE"]),
            int(item.get("DOCS_BASE_APP", 0)),
            int(item["PDFS_BRUTOS"]),
            int(item["COLETA_MD"]),
        ),
        reverse=True,
    )
    return candidates[0]


def build_report_markdown(
    overview: Dict[str, object],
    states: Sequence[Dict[str, object]],
    next_state: Dict[str, object] | None,
    e_mec_info: Dict[str, str],
) -> str:
    uf = str(overview["uf"])

    fluxo_rows = [
        ["Fonte E-MEC (arquivo original)", "1" if e_mec_info.get("exists") == "1" else "0", e_mec_info.get("path", "")],
        ["Selecao estadual (Pesquisa_" + uf + ".csv)", str(overview["reference_rows"]), str(overview["reference_csv_path"])],
        ["Prompts 7x7 gerados", str(overview["prompts_md_count"]), "prompts_individuais_7em7/" + uf],
        ["Coletas markdown", str(overview["coleta_md_count"]), "coleta_md_por_estado/" + uf],
        ["PDFs baixados", str(overview["pdf_count"]), "pdfs_baixados/" + uf],
        ["Base app (PDFs)", str(overview["docs_base"]), "Analyse_IA/app_IA/data/base_documentos.csv"],
        ["Analises IA concluidas", str(overview["docs_analisados"]), "Analyse_IA/Resposta_csv/analise_notebooklm_consolidado.csv"],
        ["Texto qualitativo (md)", str(overview["texto_qualitativo_md_count"]), "Analyse_IA/Texto_Qualitativo/" + uf],
        ["Validacao manual (md)", str(overview["validacao_manual_md_count"]), "Analyse_IA/Validação_Manual/" + uf],
    ]

    status_rows = [[k, v] for k, v in sorted((overview["validacao_status"] or {}).items())]
    comparacao_rows = [[k, v] for k, v in sorted((overview["comparacao_status"] or {}).items())]

    docs_rows = []
    for row in overview["docs_detail"]:
        doc_key = row.get("DOC_KEY", "")
        incons = (overview["inconsistencia_por_doc"] or {}).get(doc_key, {})
        docs_rows.append(
            [
                row.get("DOC_ID", ""),
                row.get("NOME_ARQUIVO_LOCAL", ""),
                row.get("STATUS_VALIDACAO_HUMANA", ""),
                row.get("STATUS_COMPARACAO_BASE", ""),
                row.get("PRESENCA_TDIC_0_1", ""),
                row.get("PRESENCA_IA_0_1", ""),
                row.get("NIVEL_INTEGRACAO_0_1_2_3", ""),
                str(incons.get("incons_ies_pct", "")) + "%" if incons else "",
            ]
        )

    states_rows = [
        [
            item["UF"],
            item["REF_ROWS"],
            item["PDFS_BRUTOS"],
            item["DOCS_BASE_APP"],
            item["ANALISADOS"],
            item["PENDENTES_ANALISE"],
            str(item["COBERTURA_ANALISE_PDFS_PCT"]) + "%",
        ]
        for item in states
    ]

    pending_docs = ", ".join(overview["docs_pending"]) if overview["docs_pending"] else "Nenhum"

    if next_state:
        docs_base_txt = (
            str(next_state.get("DOCS_BASE_APP")) if int(next_state.get("DOCS_BASE_APP", 0)) > 0 else "NA"
        )
        next_state_text = (
            f"Proximo estado recomendado: {next_state['UF']} (PDFs brutos={next_state['PDFS_BRUTOS']}, "
            f"DOCs base app={docs_base_txt}, "
            f"analises pendentes={next_state['PENDENTES_ANALISE']})."
        )
    else:
        next_state_text = "Nao foi identificado estado com PDFs prontos e analises pendentes."

    cobertura_txt = f"{overview['docs_analisados']}/{overview['docs_base']} ({overview['coverage_pct']}%)"

    if float(overview["coverage_pct"]) < 100.0:
        risco_texto = (
            f"Risco atual (se publicar agora sem completar estado): cobertura parcial "
            f"({overview['coverage_pct']}%) pode reduzir forca inferencial."
        )
    else:
        risco_texto = (
            "Risco residual: mesmo com cobertura de 100%, ainda ha ruido de vinculo IES x campus na base "
            "auxiliar e isso exige validacao humana explicita na escrita."
        )

    report = f"""
# RELATORIO GERAL DO ESTADO {uf} PARA ESCRITA DO ARTIGO

Data de geracao (UTC): {now_utc_iso()}

## 1) Escopo e trilha metodologica auditavel

Este relatorio consolida a trilha completa usada no projeto para o estado {uf}:
1. Fonte original (planilha E-MEC).
2. Recorte estadual (Pesquisa_{uf}.csv).
3. Geracao de prompts 7x7 para pesquisa profunda.
4. Coleta markdown por estado.
5. Download de PDFs.
6. Analise assistida por IA (NotebookLM via app local).
7. Validacao manual humana e consolidacao em CSV.

{markdown_table(["Etapa", "Quantidade", "Fonte/Pasta"], fluxo_rows)}

## 2) Diagnostico quantitativo de cobertura

- Cobertura de analise sobre base app: {cobertura_txt}
- Documentos pendentes de analise no estado {uf}: {pending_docs}
- Presenca TDIC (entre analisados): {overview['tdic_present']}/{overview['docs_analisados']}
- Presenca IA (entre analisados): {overview['ia_present']}/{overview['docs_analisados']}
- Nivel medio de integracao curricular (0-3): {overview['nivel_integracao_media']}
- Inconsistencia IES nos vinculos base (global): {overview['inconsistencia_vinculos_ies']}/{overview['inconsistencia_vinculos_total']} ({overview['inconsistencia_vinculos_ies_pct']}%)

### Status de validacao humana

{markdown_table(["Status", "Quantidade"], status_rows if status_rows else [["SEM_DADOS", 0]])}

### Status de comparacao com a base local

{markdown_table(["Status comparacao", "Quantidade"], comparacao_rows if comparacao_rows else [["SEM_DADOS", 0]])}

## 3) Diagnostico por documento ja analisado

{markdown_table(["DOC_ID", "Arquivo", "Validacao", "Comparacao base", "TDIC", "IA", "Integracao", "Inconsistencia IES no vinculo"], docs_rows if docs_rows else [["SEM_DADOS", "", "", "", "", "", "", ""]])}

## 4) Interpretacao tecnica das divergencias (inclui DOC0002)

Conclusao principal: as divergencias observadas ate aqui estao majoritariamente associadas a ruido da base de vinculos (instituicao/modalidade/campus), e nao a erro simples de prompt.

Indicadores objetivos:
- O app registrou alta taxa de inconsistencias IES em vinculos para {uf} ({overview['inconsistencia_vinculos_ies_pct']}%).
- O caso DOC0002 apresenta status DIVERGENTE e conflito instituicao do PDF vs base local.
- Isso reforca que a comparacao com base deve ser tratada como evidencial auxiliar e sempre mediada por validacao humana.

## 5) Viabilidade cientifica para artigo com o corpus atual

Resposta curta: SIM, e possivel produzir artigo de qualidade, com condicoes metodologicas claras.

Condicoes para manter consistencia com a realidade:
1. Enquadrar o estudo como analise documental estadual (caso {uf}) nesta fase.
2. Declarar explicitamente limitacoes de disponibilidade documental e de vinculo institucional.
3. Manter regra dura de evidencia literal + secao/pagina para todas as inferencias.
4. Diferenciar com clareza: dado do documento vs dado de base auxiliar vs decisao humana final.
5. Finalizar os documentos pendentes do estado antes de generalizacoes mais amplas.

{risco_texto}

## 6) Plano curto para fechar {uf} e iniciar proximo estado

1. Concluir pendentes de {uf}: {pending_docs}
2. Revisar casos PARCIAL/DIVERGENTE com foco em vinculo IES x campus.
3. Congelar base analitica do estado (snapshot final para escrita).
4. Exportar resultados para o notebook de redacao do artigo.

{next_state_text}

## 7) Mapa geral de estados (prontidao para analise)

{markdown_table(["UF", "Ref rows", "PDFs brutos", "DOCs base app", "Analisados", "Pendentes", "Cobertura PDFs"], states_rows)}

## 8) Pacote para o notebook de escrita

Sugestao de insumos para enviar ao notebook de redacao:
1. Este relatorio markdown.
2. CSV consolidado de analise.
3. CSV de controle de validacao manual.
4. Pastas de texto qualitativo e validacao manual do estado {uf}.

Arquivos-chave:
- Analyse_IA/Resposta_csv/analise_notebooklm_consolidado.csv
- Analyse_IA/Validação_Manual/validacao_controle.csv
- Analyse_IA/Texto_Qualitativo/{uf}/
- Analyse_IA/Validação_Manual/{uf}/
""".strip()

    return report + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera relatorio geral por estado para escrita de artigo.")
    parser.add_argument("--uf", required=True, help="UF alvo (ex.: AL)")
    parser.add_argument(
        "--e-mec-path",
        default=r"C:\Users\augus\OneDrive\Documentos\TCC\Projeto_0.2\Documentos\Metadados\Planilha_Total_info_Matemática.xlsx",
        help="Caminho da planilha original E-MEC",
    )
    parser.add_argument(
        "--output-dir",
        default="",
        help="Diretorio de saida. Padrao: new_date/Analyse_IA/Relatorios",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    script_path = Path(__file__)
    paths = discover_paths(script_path)

    uf = args.uf.upper()

    states = list_state_counts(paths)
    overview = compute_state_overview(paths, uf)
    next_state = recommend_next_state(states, current_uf=uf)

    e_mec_info = maybe_xlsx_stats(Path(args.e_mec_path))

    report_md = build_report_markdown(overview, states, next_state, e_mec_info)

    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        output_dir = paths.analyse_root / "Relatorios"

    output_dir.mkdir(parents=True, exist_ok=True)

    date_tag = datetime.now().strftime("%Y%m%d")
    md_path = output_dir / f"RELATORIO_GERAL_{uf}_{date_tag}.md"
    json_path = output_dir / f"RELATORIO_GERAL_{uf}_{date_tag}.json"

    md_path.write_text(report_md, encoding="utf-8")

    payload = {
        "generated_at_utc": now_utc_iso(),
        "uf": uf,
        "overview": overview,
        "states": states,
        "next_state": next_state,
        "e_mec_info": e_mec_info,
        "report_markdown_path": str(md_path),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Relatorio gerado: {md_path}")
    print(f"Dados estruturados: {json_path}")
    if next_state:
        print(f"Proximo estado recomendado: {next_state['UF']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
